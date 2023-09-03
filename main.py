# import openpyxl

# # Mở tệp Excel chứa 50 tên
# workbook = openpyxl.load_workbook('Book1.xlsx')
# sheet_goc = workbook.active

# # Tạo 5 sheet mới
# for i in range(1, 6):
#     new_sheet = workbook.create_sheet(title=f'Sheet{i}')

# # Tạo danh sách tên không lặp lại
# ten_khong_lap_lai = list(set(sheet_goc['A1:A50']))

# # Phân chia danh sách tên vào 5 phần bằng số phần tử gần bằng nhau
# so_phan_tu_moi_sheet = len(ten_khong_lap_lai) // 5

# # Bắt đầu phân chia tên vào các sheet
# for i, sheet in enumerate(workbook.sheetnames[1:], start=1):
#     current_sheet = workbook[sheet]
#     for j in range(so_phan_tu_moi_sheet):
#         # Chuyển tuple thành giá trị cụ thể (value)
#         ten = ten_khong_lap_lai.pop()[0].value
#         current_sheet.cell(row=j + 1, column=1, value=ten)

# # Lưu tệp Excel sau khi đã chỉnh sửa
# workbook.save('tên_mới.xlsx')

import openpyxl

# Mở tệp Excel chứa 50 tên và dữ liệu tương ứng
workbook = openpyxl.load_workbook('Book1.xlsx')
sheet_goc = workbook.active

# Tạo 5 sheet mới
for i in range(1, 6):
    new_sheet = workbook.create_sheet(title=f'To_{i}')

# Tạo danh sách các dòng không lặp lại và cột A và B tương ứng
danh_sach_dong = list(set((cell[0].value, cell[1].value) for cell in sheet_goc.iter_rows(min_row=1, max_row=50, min_col=1, max_col=2)))

# Phân chia danh sách dòng vào 5 phần bằng số dòng gần bằng nhau
so_dong_moi_sheet = len(danh_sach_dong) // 5

# Bắt đầu phân chia dữ liệu vào các sheet
for i, sheet in enumerate(workbook.sheetnames[1:], start=1):
    current_sheet = workbook[sheet]
    for dong in danh_sach_dong[:so_dong_moi_sheet]:
        current_sheet.append(dong)
    danh_sach_dong = danh_sach_dong[so_dong_moi_sheet:]

# Lưu tệp Excel sau khi đã chỉnh sửa
workbook.save('DANH_SACH_TO.xlsx')
